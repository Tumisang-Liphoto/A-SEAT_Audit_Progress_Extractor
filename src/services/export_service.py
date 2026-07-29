from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExportService:
    """Exports extracted audit records to Excel and CSV."""

    DATE_COLUMNS = [
        "Planned Start Date",
        "Planned Completion Date",
    ]

    @staticmethod
    def _create_output_directory(
        base_folder: str,
    ) -> Path:
        """Create and return the monthly output directory."""

        output_directory = Path(
            base_folder
        ).expanduser()

        if not output_directory.is_absolute():
            output_directory = (
                output_directory.resolve()
            )

        month_directory = (
            output_directory
            / datetime.now().strftime(
                "%Y-%m"
            )
        )

        month_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        return month_directory

    @staticmethod
    def _create_filename(
        extension: str,
        audit_year: str = "",
    ) -> str:
        """Create a descriptive timestamped output filename."""

        timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        safe_year = "".join(
            character
            for character in audit_year
            if character.isdigit()
        )

        year_section = (
            f"_AuditYear-{safe_year}"
            if safe_year
            else ""
        )

        return (
            "A-SEAT_Audit_Progress"
            f"{year_section}"
            f"_Extracted-{timestamp}.{extension}"
        )

    @staticmethod
    def _progress_heading() -> str:
        """Return the dated heading used for the progress column."""

        extraction_date = datetime.now().strftime(
            "%d %B %Y"
        )

        return (
            f"Progress to {extraction_date}"
        )

    @classmethod
    def _prepare_dataframe(
        cls,
        records: list[dict[str, Any]],
    ) -> pd.DataFrame:
        """Prepare extracted records for export."""

        if not records:
            raise ValueError(
                "There are no extracted records to export."
            )

        dataframe = pd.DataFrame(
            records
        )

        expected_columns = [
            "Auditee Name",
            "Directorate",
            "Audit Name",
            "Audit Lead",
            "Audit Type",
            "Planned Start Date",
            "Planned Completion Date",
            "Audit Year",
            "Progress",
        ]

        for column in expected_columns:
            if column not in dataframe.columns:
                dataframe[column] = ""

        dataframe = dataframe[
            expected_columns
        ]

        for date_column in cls.DATE_COLUMNS:
            dataframe[
                date_column
            ] = pd.to_datetime(
                dataframe[
                    date_column
                ],
                errors="coerce",
            )

        dataframe = dataframe.rename(
            columns={
                "Progress": (
                    cls._progress_heading()
                ),
            }
        )

        return dataframe

    def export_excel(
        self,
        records: list[dict[str, Any]],
        output_folder: str,
        audit_year: str = "",
    ) -> Path:
        """Export audit records to a formatted Excel workbook."""

        output_directory = (
            self._create_output_directory(
                output_folder
            )
        )

        output_path = (
            output_directory
            / self._create_filename(
                "xlsx",
                audit_year,
            )
        )

        dataframe = (
            self._prepare_dataframe(
                records
            )
        )

        progress_heading = (
            self._progress_heading()
        )

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            datetime_format="DD MMMM YYYY",
        ) as writer:
            dataframe.to_excel(
                writer,
                index=False,
                sheet_name="Audit Progress",
            )

            worksheet = writer.sheets[
                "Audit Progress"
            ]

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            header_map = {
                cell.value: cell.column
                for cell in worksheet[1]
            }

            for date_heading in self.DATE_COLUMNS:
                column_number = (
                    header_map.get(
                        date_heading
                    )
                )

                if column_number is None:
                    continue

                for row_number in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).number_format = (
                        "dd mmmm yyyy"
                    )

            progress_column = (
                header_map.get(
                    progress_heading
                )
            )

            if progress_column is not None:
                for row_number in range(
                    2,
                    worksheet.max_row + 1,
                ):
                    cell = worksheet.cell(
                        row=row_number,
                        column=progress_column,
                    )

                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )

                    if (
                        cell.value
                        and "\n" in str(
                            cell.value
                        )
                    ):
                        line_count = (
                            str(
                                cell.value
                            ).count(
                                "\n"
                            )
                            + 1
                        )

                        worksheet.row_dimensions[
                            row_number
                        ].height = max(
                            15,
                            line_count * 15,
                        )

            for column_cells in worksheet.columns:
                maximum_length = 0

                column_letter = (
                    column_cells[
                        0
                    ].column_letter
                )

                for cell in column_cells:
                    value_length = len(
                        str(
                            cell.value
                            or ""
                        )
                    )

                    maximum_length = max(
                        maximum_length,
                        value_length,
                    )

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max(
                        maximum_length + 3,
                        12,
                    ),
                    40,
                )

        return output_path

    def add_analysis_sheets(
        self,
        workbook_path: str | Path,
        comparison: dict[str, Any],
    ) -> Path:
        """
        Add management-summary and audit-analysis sheets to an
        existing Excel extraction workbook.
        """

        output_path = Path(
            workbook_path
        )

        if not output_path.is_file():
            raise FileNotFoundError(
                (
                    "The Excel workbook could not be found: "
                    f"{output_path}"
                )
            )

        if output_path.suffix.lower() != ".xlsx":
            raise ValueError(
                (
                    "Analysis sheets can only be added "
                    "to .xlsx files."
                )
            )

        workbook = load_workbook(
            output_path
        )

        for sheet_name in (
            "Analysis Summary",
            "Audit Analysis",
        ):
            if sheet_name in workbook.sheetnames:
                del workbook[
                    sheet_name
                ]

        summary_sheet = workbook.create_sheet(
            "Analysis Summary"
        )

        analysis_sheet = workbook.create_sheet(
            "Audit Analysis"
        )

        self._build_analysis_summary_sheet(
            summary_sheet,
            comparison,
        )

        self._build_audit_analysis_sheet(
            analysis_sheet,
            comparison,
        )

        workbook.save(
            output_path
        )

        return output_path

    @classmethod
    def _build_analysis_summary_sheet(
        cls,
        worksheet: Any,
        comparison: dict[str, Any],
    ) -> None:
        """Create the management summary sheet."""

        title_fill = PatternFill(
            "solid",
            fgColor="17365D",
        )

        section_fill = PatternFill(
            "solid",
            fgColor="D9EAF7",
        )

        header_fill = PatternFill(
            "solid",
            fgColor="5B9BD5",
        )

        white_font = Font(
            color="FFFFFF",
            bold=True,
        )

        bold_font = Font(
            bold=True
        )

        thin_border = Border(
            left=Side(
                style="thin",
                color="B7C9D6",
            ),
            right=Side(
                style="thin",
                color="B7C9D6",
            ),
            top=Side(
                style="thin",
                color="B7C9D6",
            ),
            bottom=Side(
                style="thin",
                color="B7C9D6",
            ),
        )

        worksheet.sheet_view.showGridLines = (
            False
        )

        worksheet.freeze_panes = "A5"

        worksheet.merge_cells(
            "A1:D1"
        )

        worksheet["A1"] = (
            "A-SEAT Audit Progress Analysis Summary"
        )

        worksheet["A1"].fill = (
            title_fill
        )

        worksheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=16,
        )

        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[
            1
        ].height = 28

        assessment_date = str(
            comparison.get(
                "assessment_display_date",
                "",
            )
        ).strip()

        previous_date = str(
            comparison.get(
                "previous_display_date",
                "",
            )
        ).strip()

        current_date = str(
            comparison.get(
                "current_display_date",
                "",
            )
        ).strip()

        worksheet["A3"] = (
            "Assessment date"
        )

        worksheet["B3"] = (
            assessment_date
        )

        worksheet["C3"] = (
            "Comparison period"
        )

        worksheet["D3"] = (
            f"{previous_date} to {current_date}"
            if previous_date or current_date
            else "Not available"
        )

        for cell in worksheet[3]:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        worksheet["A3"].font = (
            bold_font
        )

        worksheet["C3"].font = (
            bold_font
        )

        summary = comparison.get(
            "summary",
            {},
        )

        delivery_summary = comparison.get(
            "delivery_summary",
            {},
        )

        movement_rows = [
            (
                "Audits compared",
                cls._safe_number(
                    summary.get(
                        "audits_compared",
                        0,
                    )
                ),
            ),
            (
                "Progressed",
                cls._safe_number(
                    summary.get(
                        "progressed",
                        0,
                    )
                ),
            ),
            (
                "No change",
                cls._safe_number(
                    summary.get(
                        "unchanged",
                        0,
                    )
                ),
            ),
            (
                "Regressed",
                cls._safe_number(
                    summary.get(
                        "regressed",
                        0,
                    )
                ),
            ),
            (
                "New audits",
                cls._safe_number(
                    summary.get(
                        "new",
                        0,
                    )
                ),
            ),
            (
                "No longer listed",
                cls._safe_number(
                    summary.get(
                        "missing",
                        0,
                    )
                ),
            ),
            (
                "Not comparable",
                cls._safe_number(
                    summary.get(
                        "not_comparable",
                        0,
                    )
                ),
            ),
        ]

        data_issue_count = sum(
            cls._safe_number(
                delivery_summary.get(
                    key,
                    0,
                )
            )
            for key in (
                "missing_progress",
                "invalid_progress",
                "progress_year_mismatch",
                "missing_dates",
                "invalid_dates",
            )
        )

        delivery_rows = [
            (
                "Completed",
                cls._safe_number(
                    delivery_summary.get(
                        "completed",
                        0,
                    )
                ),
            ),
            (
                "Overdue",
                cls._safe_number(
                    delivery_summary.get(
                        "overdue",
                        0,
                    )
                ),
            ),
            (
                "Due soon",
                cls._safe_number(
                    delivery_summary.get(
                        "due_soon",
                        0,
                    )
                ),
            ),
            (
                "Not started late",
                cls._safe_number(
                    delivery_summary.get(
                        "not_started_late",
                        0,
                    )
                ),
            ),
            (
                "In progress",
                cls._safe_number(
                    delivery_summary.get(
                        "in_progress",
                        0,
                    )
                ),
            ),
            (
                "Not yet started",
                cls._safe_number(
                    delivery_summary.get(
                        "not_yet_started",
                        0,
                    )
                ),
            ),
            (
                "Data issues",
                data_issue_count,
            ),
        ]

        worksheet.merge_cells(
            "A5:B5"
        )

        worksheet["A5"] = (
            "Progress Movement"
        )

        worksheet["A5"].fill = (
            section_fill
        )

        worksheet["A5"].font = (
            bold_font
        )

        worksheet["A5"].alignment = Alignment(
            horizontal="center",
        )

        worksheet.merge_cells(
            "C5:D5"
        )

        worksheet["C5"] = (
            "Current Delivery Status"
        )

        worksheet["C5"].fill = (
            section_fill
        )

        worksheet["C5"].font = (
            bold_font
        )

        worksheet["C5"].alignment = Alignment(
            horizontal="center",
        )

        for column, heading in (
            (
                "A",
                "Movement",
            ),
            (
                "B",
                "Count",
            ),
            (
                "C",
                "Delivery status",
            ),
            (
                "D",
                "Count",
            ),
        ):
            cell = worksheet[
                f"{column}6"
            ]

            cell.value = heading
            cell.fill = header_fill
            cell.font = white_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.border = thin_border

        max_rows = max(
            len(
                movement_rows
            ),
            len(
                delivery_rows
            ),
        )

        for index in range(
            max_rows
        ):
            row_number = (
                7 + index
            )

            if index < len(
                movement_rows
            ):
                label, value = (
                    movement_rows[
                        index
                    ]
                )

                worksheet.cell(
                    row=row_number,
                    column=1,
                    value=label,
                )

                worksheet.cell(
                    row=row_number,
                    column=2,
                    value=value,
                )

            if index < len(
                delivery_rows
            ):
                label, value = (
                    delivery_rows[
                        index
                    ]
                )

                worksheet.cell(
                    row=row_number,
                    column=3,
                    value=label,
                )

                worksheet.cell(
                    row=row_number,
                    column=4,
                    value=value,
                )

            for column_number in range(
                1,
                5,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            worksheet.cell(
                row=row_number,
                column=2,
            ).alignment = Alignment(
                horizontal="center",
            )

            worksheet.cell(
                row=row_number,
                column=4,
            ).alignment = Alignment(
                horizontal="center",
            )

        average_movement = comparison.get(
            "average_movement",
            0,
        )

        note_row = 15

        worksheet.merge_cells(
            start_row=note_row,
            start_column=1,
            end_row=note_row,
            end_column=4,
        )

        worksheet.cell(
            row=note_row,
            column=1,
            value=(
                "Average movement across comparable audits: "
                f"{average_movement} percentage points."
            ),
        )

        worksheet.cell(
            row=note_row,
            column=1,
        ).fill = PatternFill(
            "solid",
            fgColor="E2F0D9",
        )

        worksheet.cell(
            row=note_row,
            column=1,
        ).font = bold_font

        worksheet.cell(
            row=note_row,
            column=1,
        ).alignment = Alignment(
            wrap_text=True,
        )

        movement_chart = BarChart()
        movement_chart.type = "col"
        movement_chart.style = 10

        movement_chart.title = (
            "Progress Movement"
        )

        movement_chart.y_axis.title = (
            "Number of audits"
        )

        movement_chart.x_axis.title = (
            "Movement"
        )

        movement_data = Reference(
            worksheet,
            min_col=2,
            min_row=6,
            max_row=13,
        )

        movement_categories = Reference(
            worksheet,
            min_col=1,
            min_row=7,
            max_row=13,
        )

        movement_chart.add_data(
            movement_data,
            titles_from_data=True,
        )

        movement_chart.set_categories(
            movement_categories
        )

        movement_chart.height = 7
        movement_chart.width = 12
        movement_chart.legend = None

        worksheet.add_chart(
            movement_chart,
            "F5",
        )

        delivery_chart = BarChart()
        delivery_chart.type = "col"
        delivery_chart.style = 11

        delivery_chart.title = (
            "Current Delivery Status"
        )

        delivery_chart.y_axis.title = (
            "Number of audits"
        )

        delivery_chart.x_axis.title = (
            "Delivery status"
        )

        delivery_data = Reference(
            worksheet,
            min_col=4,
            min_row=6,
            max_row=13,
        )

        delivery_categories = Reference(
            worksheet,
            min_col=3,
            min_row=7,
            max_row=13,
        )

        delivery_chart.add_data(
            delivery_data,
            titles_from_data=True,
        )

        delivery_chart.set_categories(
            delivery_categories
        )

        delivery_chart.height = 7
        delivery_chart.width = 12
        delivery_chart.legend = None

        worksheet.add_chart(
            delivery_chart,
            "F20",
        )

        worksheet.column_dimensions[
            "A"
        ].width = 24

        worksheet.column_dimensions[
            "B"
        ].width = 12

        worksheet.column_dimensions[
            "C"
        ].width = 26

        worksheet.column_dimensions[
            "D"
        ].width = 14

    @classmethod
    def _build_audit_analysis_sheet(
        cls,
        worksheet: Any,
        comparison: dict[str, Any],
    ) -> None:
        """Create the audit-level analysis sheet."""

        rows = [
            dict(
                row
            )
            for row in comparison.get(
                "rows",
                [],
            )
            if isinstance(
                row,
                dict,
            )
        ]

        headers = [
            "Auditee Name",
            "Directorate",
            "Audit Name",
            "Audit Lead",
            "Audit Type",
            "Audit Year",
            "Planned Start Date",
            "Planned Completion Date",
            "Previous Progress",
            "Current Progress",
            "Movement",
            "Movement Status",
            "Delivery Status",
            "Days to Completion",
            "Issue",
            "Match Method",
            "Match Score (%)",
            "Identity Change",
        ]

        worksheet.sheet_view.showGridLines = (
            False
        )

        worksheet.freeze_panes = "A2"

        header_fill = PatternFill(
            "solid",
            fgColor="17365D",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        thin_border = Border(
            left=Side(
                style="thin",
                color="D9E2F3",
            ),
            right=Side(
                style="thin",
                color="D9E2F3",
            ),
            top=Side(
                style="thin",
                color="D9E2F3",
            ),
            bottom=Side(
                style="thin",
                color="D9E2F3",
            ),
        )

        for (
            column_number,
            heading,
        ) in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column_number,
                value=heading,
            )

            cell.fill = header_fill
            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = thin_border

        status_fills = {
            "Completed": "C6E0B4",
            "Overdue": "F4CCCC",
            "Due soon": "FFF2CC",
            "Not started late": "FCE4D6",
            "In progress": "D9EAF7",
            "Not yet started": "EDEDED",
            "Progress-year mismatch": "E4DFEC",
            "Missing progress": "E4DFEC",
            "Invalid progress": "E4DFEC",
            "Missing dates": "E4DFEC",
            "Invalid dates": "E4DFEC",
            "Not currently listed": "D9D9D9",
        }

        for (
            row_number,
            row,
        ) in enumerate(
            rows,
            start=2,
        ):
            values = [
                row.get(
                    "Auditee Name",
                    "",
                ),
                row.get(
                    "Directorate",
                    "",
                ),
                row.get(
                    "Audit Name",
                    "",
                ),
                row.get(
                    "Audit Lead",
                    "",
                ),
                row.get(
                    "Audit Type",
                    "",
                ),
                row.get(
                    "Audit Year",
                    "",
                ),
                row.get(
                    "Planned Start Date",
                    "",
                ),
                row.get(
                    "Planned Completion Date",
                    "",
                ),
                row.get(
                    "previous_progress",
                    "",
                ),
                row.get(
                    "current_progress",
                    "",
                ),
                row.get(
                    "movement_text",
                    "",
                ),
                row.get(
                    "status_label",
                    "",
                ),
                row.get(
                    "delivery_status_label",
                    "",
                ),
                row.get(
                    "days_to_completion",
                    None,
                ),
                row.get(
                    "delivery_issue",
                    "",
                ),
                row.get(
                    "match_method_label",
                    "",
                ),
                row.get(
                    "match_score",
                    None,
                ),
                row.get(
                    "identity_change_text",
                    "",
                ),
            ]

            for (
                column_number,
                value,
            ) in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                    value=value,
                )

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            delivery_label = str(
                row.get(
                    "delivery_status_label",
                    "",
                )
            )

            delivery_fill = (
                status_fills.get(
                    delivery_label
                )
            )

            if delivery_fill:
                worksheet.cell(
                    row=row_number,
                    column=13,
                ).fill = PatternFill(
                    "solid",
                    fgColor=delivery_fill,
                )

            movement_label = str(
                row.get(
                    "status_label",
                    "",
                )
            )

            if movement_label == "Regressed":
                worksheet.cell(
                    row=row_number,
                    column=12,
                ).fill = PatternFill(
                    "solid",
                    fgColor="F4CCCC",
                )

            elif movement_label == "Progressed":
                worksheet.cell(
                    row=row_number,
                    column=12,
                ).fill = PatternFill(
                    "solid",
                    fgColor="C6E0B4",
                )

            elif movement_label == (
                "Not comparable"
            ):
                worksheet.cell(
                    row=row_number,
                    column=12,
                ).fill = PatternFill(
                    "solid",
                    fgColor="E4DFEC",
                )

            match_method = str(
                row.get(
                    "match_method",
                    "",
                )
            )

            if match_method == "fuzzy":
                worksheet.cell(
                    row=row_number,
                    column=16,
                ).fill = PatternFill(
                    "solid",
                    fgColor="FFF2CC",
                )

            elif match_method == (
                "strong_identity"
            ):
                worksheet.cell(
                    row=row_number,
                    column=16,
                ).fill = PatternFill(
                    "solid",
                    fgColor="D9EAF7",
                )

            if bool(
                row.get(
                    "identity_changed",
                    False,
                )
            ):
                worksheet.cell(
                    row=row_number,
                    column=18,
                ).fill = PatternFill(
                    "solid",
                    fgColor="FCE4D6",
                )

        worksheet.auto_filter.ref = (
            f"A1:R{max(1, len(rows) + 1)}"
        )

        widths = {
            "A": 24,
            "B": 18,
            "C": 30,
            "D": 24,
            "E": 16,
            "F": 12,
            "G": 19,
            "H": 22,
            "I": 18,
            "J": 18,
            "K": 22,
            "L": 18,
            "M": 23,
            "N": 18,
            "O": 48,
            "P": 27,
            "Q": 17,
            "R": 55,
        }

        for (
            column_letter,
            width,
        ) in widths.items():
            worksheet.column_dimensions[
                column_letter
            ].width = width

        worksheet.row_dimensions[
            1
        ].height = 32

    @staticmethod
    def _safe_number(
        value: Any,
    ) -> int:
        """Convert a summary value to an integer."""

        try:
            return int(
                value
            )
        except (
            TypeError,
            ValueError,
        ):
            return 0

    def export_csv(
        self,
        records: list[dict[str, Any]],
        output_folder: str,
        audit_year: str = "",
    ) -> Path:
        """Export audit records to CSV."""

        output_directory = (
            self._create_output_directory(
                output_folder
            )
        )

        output_path = (
            output_directory
            / self._create_filename(
                "csv",
                audit_year,
            )
        )

        dataframe = (
            self._prepare_dataframe(
                records
            )
        )

        for date_column in self.DATE_COLUMNS:
            dataframe[
                date_column
            ] = (
                dataframe[
                    date_column
                ]
                .dt.strftime(
                    "%d %B %Y"
                )
                .fillna(
                    ""
                )
            )

        dataframe.to_csv(
            output_path,
            index=False,
            encoding="utf-8-sig",
        )

        return output_path